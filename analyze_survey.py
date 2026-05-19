import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('Trivia Platform Quality Assurance (QA) Survey (Responses).xlsx')
ws = wb.active

# Parse all responses
responses = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    responses.append({
        'timestamp': row[0],
        'platform': row[1] or 'Unknown',
        'accuracy_rating': row[2],  # 1-5 scale
        'clarity': row[3],
        'timer': row[4],
        'difficulty': row[5],
        'errors_feedback': row[6]
    })

total = len(responses)
print(f"{'='*60}")
print(f"  SURVEY ANALYSIS REPORT ({total} Responses)")
print(f"{'='*60}")

# Platform breakdown
platforms = Counter(r['platform'] for r in responses)
print(f"\n--- PLATFORM DISTRIBUTION ---")
for p, c in platforms.most_common():
    print(f"  {p}: {c} ({c/total*100:.0f}%)")

# Accuracy ratings
print(f"\n--- FACTUAL ACCURACY RATINGS (1=Low, 5=High) ---")
for platform in ['Cashrush', 'MindQuest', 'Unknown']:
    subset = [r for r in responses if r['platform'] == platform]
    if not subset:
        continue
    ratings = [r['accuracy_rating'] for r in subset if r['accuracy_rating'] is not None]
    if ratings:
        avg = sum(ratings) / len(ratings)
        dist = Counter(ratings)
        print(f"  {platform} (n={len(ratings)}): avg={avg:.1f}/5")
        for score in sorted(dist.keys()):
            print(f"    Rating {int(score)}: {dist[score]} responses")

# Overall accuracy
all_ratings = [r['accuracy_rating'] for r in responses if r['accuracy_rating'] is not None]
print(f"  OVERALL (n={len(all_ratings)}): avg={sum(all_ratings)/len(all_ratings):.1f}/5")

# Clarity
print(f"\n--- QUESTION CLARITY ---")
clarity_counts = Counter(r['clarity'] for r in responses)
for c, count in clarity_counts.most_common():
    print(f"  {c}: {count} ({count/total*100:.0f}%)")

# By platform
for platform in ['Cashrush', 'MindQuest']:
    subset = [r for r in responses if r['platform'] == platform]
    if not subset:
        continue
    cl = Counter(r['clarity'] for r in subset)
    print(f"  [{platform}]:")
    for c, count in cl.most_common():
        print(f"    {c}: {count}")

# Timer
print(f"\n--- TIMER APPROPRIATENESS ---")
timer_counts = Counter(r['timer'] for r in responses)
for t, count in timer_counts.most_common():
    print(f"  {t}: {count} ({count/total*100:.0f}%)")

# By platform
for platform in ['Cashrush', 'MindQuest']:
    subset = [r for r in responses if r['platform'] == platform]
    if not subset:
        continue
    tc = Counter(r['timer'] for r in subset)
    print(f"  [{platform}]:")
    for t, count in tc.most_common():
        print(f"    {t}: {count}")

# Difficulty
print(f"\n--- DIFFICULTY RATING ---")
diff_counts = Counter(r['difficulty'] for r in responses)
for d, count in diff_counts.most_common():
    print(f"  {d}: {count} ({count/total*100:.0f}%)")

# By platform
for platform in ['Cashrush', 'MindQuest']:
    subset = [r for r in responses if r['platform'] == platform]
    if not subset:
        continue
    dc = Counter(r['difficulty'] for r in subset)
    print(f"  [{platform}]:")
    for d, count in dc.most_common():
        print(f"    {d}: {count}")

# Error feedback
print(f"\n--- ERROR/TYPO REPORTS ---")
no_error_keywords = ['no', 'none', 'n/a', 'nil', 'ni', "didn't see"]
meaningful_feedback = []
for r in responses:
    fb = (r['errors_feedback'] or '').strip()
    if fb and not any(fb.lower().startswith(k) or fb.lower() == k for k in no_error_keywords):
        if len(fb) > 10:  # Filter out very short non-answers
            meaningful_feedback.append((r['platform'], fb))

print(f"  Meaningful feedback entries: {len(meaningful_feedback)}/{total}")
for platform, fb in meaningful_feedback:
    print(f"  [{platform}]: {fb}")
